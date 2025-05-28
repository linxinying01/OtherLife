import logging
import os
import openpyxl
from xml.etree import ElementTree as ET
from xml.dom import minidom

logging.basicConfig(level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')


def _open_excel(excel_path):
    """加载Excel工作簿"""
    try:
        wb = openpyxl.load_workbook(excel_path)
    except FileNotFoundError:
        raise FileNotFoundError(f"错误：文件 '{excel_path}' 未找到")
    except Exception as e:
        raise Exception(f"打开Excel文件时出错: {str(e)}")

    return wb


def convert_all_excel(excel_dir, xml_dir):
    """
    将指定目录下所有excel所有sheet转换为XML文件
    """
    logging.debug("excel_dir: %s", excel_dir)
    logging.debug("xml_dir: %s", xml_dir)
    
    if not os.path.exists(excel_dir):
        raise FileNotFoundError(f"错误：excel目录 '{excel_dir}' 不存在")
    
    for root, _, files in os.walk(excel_dir):
        for file in files:
            if file.endswith('.xlsx') or file.endswith('.xlsm'):
                excel_path = os.path.join(root, file)
                convert_one_excel(excel_path, xml_dir)


def convert_one_excel(excel_path, xml_dir):
    """
    将指定Excel工作表的所有sheet转换为XML文件
    """
    logging.debug("excel_path: %s", excel_path)
    logging.debug("xml_dir: %s", xml_dir)
    
    wb = _open_excel(excel_path)
    
    for sheet_name in wb.sheetnames:
        convert_one_sheet(excel_path, sheet_name, xml_dir)
    

def convert_one_sheet(excel_path, sheet_name, xml_dir):
    """
    将指定Excel工作表的某个sheet转换为XML文件
    """
    logging.debug("excel_path: %s", excel_path)
    logging.debug("xml_dir: %s", xml_dir)
    logging.debug("sheet_name: %s", sheet_name)
    
    wb = _open_excel(excel_path)
    
    if sheet_name is None or sheet_name == '':
        sheet_name = wb.sheetnames[0]  # 默认使用第一个工作表
        
    if sheet_name not in wb.sheetnames:
        logging.warning("警告：工作簿 '%s' 中没有名为 '%s' 的工作表，跳过转换", excel_path, sheet_name)
        return
    
    sheet = wb[sheet_name]
    
    # 创建XML根元素
    root = ET.Element("root")
    
    xml_name = sheet[1][0].value # 第一行第一列是XML文件名
    if xml_name is None or xml_name == '' or not xml_name.strip().endswith('.xml'):
        logging.warning("警告：工作表 '%s' 的 '%s' 第一行第一列 xml_name 不能为空", excel_path, sheet_name)
        return
    
    logging.debug("XML文件名: %s", xml_name)
    
    headers = []
    # 获取列标题（假设第二行是标题）
    for cell in sheet[3]:
        if cell.value is None or cell.value.strip() == '':
            break
        headers.append(cell.value.strip())
    
    if len(headers) == 0:
        logging.warning("警告：工作表 '%s' 的 '%s' 标题行为空，跳过转换", excel_path, sheet_name)
        return
    
    is_exist_data = False
    for row in sheet.iter_rows(min_row=4):
        data = ET.SubElement(root, "data")
        is_exist_data = True
        # logging.debug("处理行: %s", [cell.value for cell in row])2
        for idx, cell in enumerate(row):
            if idx >= len(headers):
                break
            if cell.value is None:
                continue
            header = headers[idx]
            field = ET.SubElement(data, header)
            field.text = str(cell.value) if cell.value is not None else ""
    
    if not is_exist_data:
        logging.warning("警告：工作表 '%s' 的 '%s' 无有效数据，跳过转换", excel_path, sheet_name)
        return
        
    # 生成XML字符串
    xml_str = ET.tostring(root, encoding='utf-8')

    # 格式化XML
    dom = minidom.parseString(xml_str)
    pretty_xml = dom.toprettyxml(indent="  ", encoding='utf-8')

    # 保存到文件
    xml_path = os.path.join(xml_dir, xml_name)
    # logging.debug("保存XML文件到: %s", xml_path)
    with open(xml_path, 'wb') as f:
        f.write(pretty_xml)
    logging.debug("XML文件已保存到: %s", xml_path)
    

if __name__ == "__main__":
    # 在这里修改文件路径
    excel_dir = "E:\\code\\pyprojects\\OtherLife\\data\\excel"
    excel_file = os.path.join(excel_dir, "D道具表.xlsx")
    xml_dir = "E:\\code\\pyprojects\\OtherLife\\data\\xml"
    
    # convert_one_sheet(excel_file, "", xml_dir)
    # convert_one_excel(excel_file, xml_dir)
    convert_all_excel(excel_dir, xml_dir)